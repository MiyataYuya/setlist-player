"""シートを唯一の正として app_*.csv を生成する。

シート（XLSX）の URL 列ハイパーリンク (https://youtu.be/...&t=Ns) から
video_id と start_seconds を取り出し、曲名・アーティストとあわせて
app_songs_enriched.csv / app_performances.csv / app_videos.csv を出力する。

既存の fetch_setlists / parse_setlists / build_app_data / enrich_songs パイプラインは
コメント解析の不確実性を含むため、本スクリプトに置き換える。

入力:
- scraper/sheet_source.xlsx (引数 --sheet で指定、デフォルト同梱パス)
- data/app_videos.csv (動画情報キャッシュ、無くてもよい)
- 環境変数 YOUTUBE_API_KEY (未取得動画の情報補完に使用)

出力:
- data/app_songs.csv
- data/app_songs_enriched.csv
- data/app_performances.csv
- data/app_videos.csv
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import defaultdict
from typing import Optional

import openpyxl

_DIR = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR = os.path.join(os.path.dirname(_DIR), "data")

DEFAULT_SHEET = os.path.join(_DIR, "sheet_source.xlsx")

OUT_SONGS = os.path.join(_DATA_DIR, "app_songs.csv")
OUT_SONGS_ENRICHED = os.path.join(_DATA_DIR, "app_songs_enriched.csv")
OUT_PERFORMANCES = os.path.join(_DATA_DIR, "app_performances.csv")
OUT_VIDEOS = os.path.join(_DATA_DIR, "app_videos.csv")
# 公開用 CSV (README からリンク、人間可読フォーマット)
OUT_PUBLIC_PERFORMANCES = os.path.join(_DATA_DIR, "song_performances.csv")

# シート列のインデックス (1-origin)
COL_DATE = 1
COL_NO = 2
COL_TITLE = 3
COL_ARTIST = 4
COL_COUNT = 5
COL_MODE = 6
COL_URL = 7

HEADER_ROW = 5  # 5行目が「日付/No./曲名/アーティスト/回目/(空)/URL」のヘッダ
DATA_START_ROW = 6

URL_RE = re.compile(r"(?:youtu\.be/|watch\?v=)([\w-]{11})(?:.*?[?&]t=(\d+)s?)?")
END_SECONDS_FALLBACK = 300  # 動画末尾曲の end 補完: start + 5分


def _seconds_to_hms(sec: int) -> str:
    """秒数を HH:MM:SS 形式に変換。"""
    sec = max(0, int(sec))
    h, rem = divmod(sec, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def extract_sheet_rows(xlsx_path: str) -> list[dict]:
    """シート全行から performance データを抽出。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: list[dict] = []
    skipped = 0
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        date_v = ws.cell(row=row_idx, column=COL_DATE).value
        title_v = ws.cell(row=row_idx, column=COL_TITLE).value
        no_v = ws.cell(row=row_idx, column=COL_NO).value
        if not date_v or not title_v:
            continue

        url_cell = ws.cell(row=row_idx, column=COL_URL)
        href = url_cell.hyperlink.target if url_cell.hyperlink else None
        if not href:
            skipped += 1
            continue
        m = URL_RE.search(href)
        if not m:
            skipped += 1
            continue
        video_id = m.group(1)
        start_seconds = int(m.group(2)) if m.group(2) else 0

        artist_v = ws.cell(row=row_idx, column=COL_ARTIST).value
        try:
            no_int = int(no_v) if no_v is not None else 0
        except (TypeError, ValueError):
            no_int = 0

        rows.append({
            "date": str(date_v),
            "no": no_int,
            "title": str(title_v).strip(),
            "artist": (str(artist_v).strip() if artist_v else ""),
            "video_id": video_id,
            "start_seconds": start_seconds,
            "video_title_text": str(url_cell.value or "").strip(),
        })

    if skipped:
        print(f"WARN: ハイパーリンク無し/パース失敗で {skipped} 行スキップ", file=sys.stderr)
    return rows


def assign_song_ids(perf_rows: list[dict]) -> tuple[dict[tuple[str, str], str], list[dict]]:
    """(title, artist) のペアを song_id にマップ。

    出現順に song_0001, song_0002, ... と採番。
    """
    song_id_map: dict[tuple[str, str], str] = {}
    songs: list[dict] = []
    for r in perf_rows:
        key = (r["title"], r["artist"])
        if key not in song_id_map:
            sid = f"song_{len(song_id_map) + 1:04d}"
            song_id_map[key] = sid
            songs.append({
                "song_id": sid,
                "title": r["title"],
                "artist": r["artist"],
            })
    return song_id_map, songs


def fill_end_seconds(perfs: list[dict]) -> None:
    """同一 video 内で start_seconds 昇順に並べ、次曲の start を end とする。"""
    by_video: dict[str, list[dict]] = defaultdict(list)
    for p in perfs:
        by_video[p["video_id"]].append(p)

    for vid, group in by_video.items():
        group.sort(key=lambda r: r["start_seconds"])
        for i, p in enumerate(group):
            if i + 1 < len(group):
                next_start = group[i + 1]["start_seconds"]
                # 次曲が同じ秒数や逆順の場合は +300 にフォールバック
                if next_start > p["start_seconds"]:
                    p["end_seconds"] = next_start
                else:
                    p["end_seconds"] = p["start_seconds"] + END_SECONDS_FALLBACK
            else:
                p["end_seconds"] = p["start_seconds"] + END_SECONDS_FALLBACK


VIDEO_FIELDS = ["published_at", "title", "channel_title", "duration", "view_count", "like_count"]


def load_video_cache(path: str) -> dict[str, dict]:
    cache: dict[str, dict] = {}
    if not os.path.exists(path):
        return cache
    with open(path, "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            vid = row.get("video_id")
            if not vid:
                continue
            cache[vid] = {k: row.get(k, "") for k in VIDEO_FIELDS}
    return cache


def fetch_missing_video_info(missing: list[str]) -> dict[str, dict]:
    """YouTube API で動画情報を取得。失敗時は空 dict。"""
    if not missing:
        return {}
    api_key = os.environ.get("YOUTUBE_API_KEY", "")
    if not api_key:
        print(f"WARN: YOUTUBE_API_KEY 未設定。{len(missing)}件の動画情報を取得できません", file=sys.stderr)
        return {}
    try:
        from fetch_setlists import fetch_video_info, youtube_client
    except ImportError as e:
        print(f"WARN: fetch_setlists インポート失敗: {e}", file=sys.stderr)
        return {}

    youtube = youtube_client(api_key)
    info = fetch_video_info(youtube, missing)
    print(f"YouTube API から {len(info)} 件取得", file=sys.stderr)
    out: dict[str, dict] = {}
    for vid, data in info.items():
        out[vid] = {
            "published_at": data.get("published_at") or "",
            "title": data.get("title") or "",
            "channel_title": data.get("channel_title") or "",
            "duration": data.get("duration") or "",
            "view_count": str(data.get("view_count") or ""),
            "like_count": str(data.get("like_count") or ""),
        }
    return out


def build(xlsx_path: str) -> None:
    print(f"Reading sheet: {xlsx_path}")
    perf_rows = extract_sheet_rows(xlsx_path)
    print(f"  {len(perf_rows)} performance rows extracted")

    # song_id 採番
    song_id_map, songs = assign_song_ids(perf_rows)
    print(f"  {len(songs)} unique songs")

    # video master
    video_ids = sorted({r["video_id"] for r in perf_rows})
    print(f"  {len(video_ids)} unique videos")

    # video info: キャッシュ + 不足分を API 取得
    cache = load_video_cache(OUT_VIDEOS)
    print(f"  Video cache: {len(cache)} entries")
    missing = [v for v in video_ids if v not in cache or not cache[v].get("title")]
    if missing:
        print(f"  Fetching {len(missing)} missing videos from YouTube API...")
        fetched = fetch_missing_video_info(missing)
        for vid, data in fetched.items():
            cache[vid] = data
    still_missing = [v for v in video_ids if v not in cache or not cache[v].get("title")]
    if still_missing:
        # 削除/非公開動画は対応する performance ごと除外する
        # (CLAUDE.md「配信者が動画を削除・非公開にした場合、それを回避する手段を提供しない」)
        print(f"INFO: title 未取得動画 {len(still_missing)}件を除外: {still_missing}", file=sys.stderr)
        excluded = set(still_missing)
        perf_rows = [r for r in perf_rows if r["video_id"] not in excluded]
        video_ids = [v for v in video_ids if v not in excluded]
        # 除外で参照されなくなった曲は songs からも削除
        used_song_keys = {(r["title"], r["artist"]) for r in perf_rows}
        songs = [s for s in songs if (s["title"], s["artist"]) in used_song_keys]
        # song_id を振り直し (歯抜けを避ける)
        song_id_map = {}
        renumbered_songs = []
        for s in songs:
            sid = f"song_{len(song_id_map) + 1:04d}"
            song_id_map[(s["title"], s["artist"])] = sid
            renumbered_songs.append({"song_id": sid, "title": s["title"], "artist": s["artist"]})
        songs = renumbered_songs
        print(f"  除外後: songs={len(songs)}, perfs={len(perf_rows)}, videos={len(video_ids)}", file=sys.stderr)

    # performance 採番 + end_seconds 補完
    perfs: list[dict] = []
    # 入力順を保つ (date, no 順) - extract時の順序
    perf_rows_sorted = sorted(perf_rows, key=lambda r: (r["date"], r["no"]))
    for i, r in enumerate(perf_rows_sorted, start=1):
        sid = song_id_map[(r["title"], r["artist"])]
        published_at = cache.get(r["video_id"], {}).get("published_at", "")
        perfs.append({
            "performance_id": f"perf_{i:04d}",
            "song_id": sid,
            "video_id": r["video_id"],
            "start_seconds": r["start_seconds"],
            "end_seconds": 0,
            "published_at": published_at,
        })

    fill_end_seconds(perfs)

    # CSV 出力
    os.makedirs(_DATA_DIR, exist_ok=True)

    with open(OUT_SONGS, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["song_id", "title", "artist"])
        w.writeheader()
        w.writerows(songs)

    with open(OUT_SONGS_ENRICHED, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["song_id", "title", "artist", "performance_note"])
        w.writeheader()
        for s in songs:
            w.writerow({**s, "performance_note": ""})

    with open(OUT_PERFORMANCES, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "performance_id", "song_id", "video_id",
            "start_seconds", "end_seconds", "published_at",
        ])
        w.writeheader()
        w.writerows(perfs)

    with open(OUT_VIDEOS, "w", newline="", encoding="utf-8") as f:
        fields = ["video_id"] + VIDEO_FIELDS
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for vid in video_ids:
            row = {"video_id": vid}
            row.update(cache.get(vid, {k: "" for k in VIDEO_FIELDS}))
            w.writerow(row)

    # 公開用CSV: song_performances.csv (HH:MM:SS フォーマット、video_url 付き)
    song_by_id = {s["song_id"]: s for s in songs}
    with open(OUT_PUBLIC_PERFORMANCES, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "song_name", "artist", "video_id", "video_url",
            "published_at", "start_time", "end_time",
        ])
        w.writeheader()
        for p in perfs:
            s = song_by_id[p["song_id"]]
            w.writerow({
                "song_name": s["title"],
                "artist": s["artist"],
                "video_id": p["video_id"],
                "video_url": f"https://youtu.be/{p['video_id']}?t={p['start_seconds']}",
                "published_at": p["published_at"],
                "start_time": _seconds_to_hms(p["start_seconds"]),
                "end_time": _seconds_to_hms(p["end_seconds"]),
            })

    print()
    print(f"Songs:        {len(songs):4d} -> {OUT_SONGS}")
    print(f"Songs(enrich):{len(songs):4d} -> {OUT_SONGS_ENRICHED}")
    print(f"Performances: {len(perfs):4d} -> {OUT_PERFORMANCES}")
    print(f"Videos:       {len(video_ids):4d} -> {OUT_VIDEOS}")
    print(f"Public CSV:   {len(perfs):4d} -> {OUT_PUBLIC_PERFORMANCES}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="入力 xlsx パス")
    args = parser.parse_args()
    build(args.sheet)


if __name__ == "__main__":
    main()
